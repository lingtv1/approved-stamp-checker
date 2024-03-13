
import os
from pathlib import Path
from pdf2image import convert_from_path
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
import openpyxl  # For Excel manipulation


def check_for_approval_stamp(pdf_path, poppler_path):
    try:
        # Convert only the first page to an image
        image = convert_from_path(pdf_path, poppler_path=poppler_path, first_page=1, last_page=1)[0]
    except Exception as e:
        raise ValueError(f"Failed to convert {pdf_path} to image: {e}")

    image = image.convert('L')
    image = ImageEnhance.Contrast(image).enhance(2)
    image = image.filter(ImageFilter.SHARPEN)

    text = pytesseract.image_to_string(image, lang='eng', config='--psm 11')

    if "APPROVED" in text.upper():
        return True

    return False


def list_approval_stamps(directory, poppler_path):
    results = []  # List to store results

    for entry in os.scandir(directory):
        if entry.is_file() and entry.name.lower().endswith('.pdf'):
            pdf_path = entry.path
            print(f"Checking {entry.name} for approval stamp...")

            try:
                has_stamp = check_for_approval_stamp(pdf_path, poppler_path)
            except ValueError as e:
                print(f"Error processing {entry.name}: {e}")
                continue  # Skip to the next file

            if has_stamp:
                print(f"Approved Stamp found on file ({entry.name})")
                results.append({'File Name': entry.name, 'Has Stamp': 'Yes'})
            else:
                print(f"No approval stamp found in {entry.name}")
                results.append({'File Name': entry.name, 'Has Stamp': 'No'})

    # Create DataFrame and save to Excel
    df = pd.DataFrame(results)
    df.to_excel(r"d:\stamp.xlsx", index=False)

    # Open the Excel file and highlight "No" results
    wb = openpyxl.load_workbook(r"d:\stamp.xlsx")
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming headers in row 1)
        if sheet.cell(row, 2).value == "No":  # Check if 'Has Stamp' column is "No"
            sheet.cell(row, 1).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FF98FB98')  # Green fill for 'File Name'
            sheet.cell(row, 2).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FF98FB98')  # Green fill for 'Has Stamp'

    wb.save(r"d:\stamp.xlsx")  # Save the changes

    # Open the Excel file (optional)
    os.startfile(r"d:\stamp.xlsx")  # Uncomment this line to open the file
    # os.startfile(r"d:\stamp.xlsx")  # Uncomment if you want to open the file automatically


# Set your poppler_path according to your system
poppler_path = Path(r"C:\Program Files\poppler-24.02.0\Library\bin")

# Get user input for the last part of the directory
folder_name = input("Enter the folder name to search: ")

# Construct the full directory path
base_directory = Path(r"D:\OneDrive - AMQ Solutions, Inc\Marcus' project\2020 Freight Module\Freight Invoice 2022")
directory = base_directory / folder_name

# Check if the directory exists
if not directory.exists():
    print(f"Error: Directory '{directory}' does not exist.")
else:
    list_approval_stamps(directory, poppler_path)